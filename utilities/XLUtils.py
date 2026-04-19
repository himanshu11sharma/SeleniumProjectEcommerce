import openpyxl
from openpyxl.styles import PatternFill
from pygments.lexer import words


def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(rownum,columnno).value

def writeDate(file,sheetName,rownum,columno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(rownum,columno).value=data
    workbook.save(file)

def fillGreenColor(file,sheetName,rowno,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    greenFill=PatternFill(start_color='60b212',
                          end_color='60b212',
                          fill_type='solid')
    sheet.cell(rowno,columnno).fill=greenFill
    workbook.save(file)

def fillRedColor(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000',
                       end_color='ff0000',
                       fill_type='solid')
    sheet.cell(rownum,columnno).fill=redFill
    workbook.save(file)